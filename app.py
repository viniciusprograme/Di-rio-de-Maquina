import streamlit as st
import pandas as pd
from datetime import datetime, date, time
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import io

# ─────────────────────────────────────────────
#  CONFIGURAÇÃO DA PÁGINA
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="Diário de Bordo — Máquinas",
    page_icon="⚙️",
    layout="centered",
)

# ─────────────────────────────────────────────
#  ESTILO VISUAL
# ─────────────────────────────────────────────
st.markdown("""
<style>
  @import url('https://fonts.googleapis.com/css2?family=Syne:wght@700;800&family=DM+Sans:wght@400;500&display=swap');

  html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }

  .stApp { background: #111820; color: #dde6f0; }

  .header-box {
    background: linear-gradient(135deg, #1565c0 0%, #0288d1 100%);
    border-radius: 16px;
    padding: 30px 28px 22px;
    margin-bottom: 26px;
    box-shadow: 0 8px 32px rgba(21,101,192,0.4);
  }
  .header-box h1 {
    font-family: 'Syne', sans-serif;
    font-size: 1.9rem;
    font-weight: 800;
    color: #fff;
    margin: 0 0 4px;
  }
  .header-box p { color: rgba(255,255,255,0.8); margin: 0; font-size: 0.92rem; }

  .form-card {
    background: #1b2533;
    border-radius: 14px;
    padding: 24px 20px;
    margin-bottom: 18px;
    border: 1px solid rgba(255,255,255,0.07);
  }
  .section-title {
    font-family: 'Syne', sans-serif;
    font-size: 0.72rem;
    font-weight: 700;
    letter-spacing: 2px;
    text-transform: uppercase;
    color: #29b6f6;
    margin-bottom: 14px;
  }
  .calc-box {
    background: rgba(41,182,246,0.08);
    border: 1px solid rgba(41,182,246,0.25);
    border-radius: 10px;
    padding: 14px 16px;
    margin-top: 12px;
    display: flex;
    gap: 28px;
  }
  .calc-item { text-align: center; }
  .calc-label { font-size: 0.7rem; color: #29b6f6; text-transform: uppercase; letter-spacing: 1px; }
  .calc-value { font-family: 'Syne', sans-serif; font-size: 1.4rem; color: #fff; font-weight: 700; }

  .stButton > button[kind="primary"] {
    background: linear-gradient(135deg, #1565c0, #0288d1) !important;
    color: white !important;
    border: none !important;
    border-radius: 10px !important;
    padding: 14px 0 !important;
    font-family: 'Syne', sans-serif !important;
    font-weight: 700 !important;
    font-size: 1rem !important;
    width: 100% !important;
    box-shadow: 0 4px 20px rgba(21,101,192,0.4) !important;
  }
  .stButton > button[kind="primary"]:hover {
    transform: translateY(-2px) !important;
    box-shadow: 0 8px 28px rgba(21,101,192,0.55) !important;
  }

  [data-testid="metric-container"] {
    background: #1b2533;
    border-radius: 12px;
    padding: 14px !important;
    border: 1px solid rgba(255,255,255,0.07);
  }
  [data-testid="metric-container"] label {
    color: #29b6f6 !important;
    font-size: 0.68rem !important;
    font-weight: 700 !important;
    letter-spacing: 1.5px !important;
    text-transform: uppercase !important;
  }
  [data-testid="metric-container"] [data-testid="metric-value"] {
    color: #dde6f0 !important;
    font-family: 'Syne', sans-serif !important;
    font-size: 1.5rem !important;
  }

  .stTabs [data-baseweb="tab-list"] {
    background: #1b2533;
    border-radius: 10px;
    padding: 4px;
    gap: 4px;
  }
  .stTabs [data-baseweb="tab"] {
    border-radius: 8px !important;
    color: #7a8fa6 !important;
    font-weight: 500 !important;
  }
  .stTabs [aria-selected="true"] {
    background: #1565c0 !important;
    color: white !important;
  }

  .stTextInput > div > div > input,
  .stNumberInput > div > div > input,
  .stTextArea > div > div > textarea,
  .stSelectbox > div > div > div,
  .stDateInput > div > div > input,
  .stTimeInput > div > div > input {
    background: #111820 !important;
    border: 1px solid rgba(255,255,255,0.12) !important;
    border-radius: 8px !important;
    color: #dde6f0 !important;
  }

  .footer {
    text-align: center;
    color: #3d5166;
    font-size: 0.78rem;
    margin-top: 40px;
    padding-top: 20px;
    border-top: 1px solid rgba(255,255,255,0.05);
  }
  #MainMenu, footer, header { visibility: hidden; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
#  LISTAS DE EQUIPAMENTOS
# ─────────────────────────────────────────────
EQUIPAMENTOS = [
    "KAL 02", "KAL 03", "KAL 04", "KAL 05", "KAL 06", "KAL 07", "KAL 08",
    "RS 10", "RS 11", "RS 12", "RS 14", "RS 15",
    "EP 11", "EP 12", "EP 14", "EP 15", "EP 16", "EP 17", "EP 18"
]

# ─────────────────────────────────────────────
#  ARQUIVO DE DADOS
# ─────────────────────────────────────────────
ARQUIVO_CSV = "registros_maquinas.csv"
COLUNAS = [
    "ID", "Equipamento", "Horimetro_Ultimo_Abastec",
    "Data", "Operador", "N_OS", "Cliente",
    "Hora_Inicial", "Hora_Final",
    "Horimetro_Inicial", "Horimetro_Final",
    "Hor_Trab_Equip", "Abast_Horimetro", "Media",
    "Observacoes", "Registro_Em"
]

def carregar_dados():
    if os.path.exists(ARQUIVO_CSV):
        return pd.read_csv(ARQUIVO_CSV)
    return pd.DataFrame(columns=COLUNAS)

def salvar_registro(dados: dict):
    df = carregar_dados()
    novo = pd.DataFrame([dados])
    df = pd.concat([df, novo], ignore_index=True)
    df.to_csv(ARQUIVO_CSV, index=False)

def calcular_horas(h_ini: time, h_fim: time) -> float:
    ini = h_ini.hour * 60 + h_ini.minute
    fim = h_fim.hour * 60 + h_fim.minute
    diff = fim - ini
    if diff < 0:
        diff += 24 * 60
    return round(diff / 60, 2)

def gerar_excel(df_filtrado: pd.DataFrame) -> bytes:
    """Gera planilha Excel no formato do diário original."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Diário de Bordo"

    # Verificar se o dataframe está vazio
    if df_filtrado.empty:
        ws.cell(row=1, column=1, value="Nenhum registro para exportar.")
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    # ── Estilo base ──
    header_fill  = PatternFill("solid", fgColor="1565C0")
    subhead_fill = PatternFill("solid", fgColor="0D47A1")
    alt_fill     = PatternFill("solid", fgColor="1B2533")
    white_font   = Font(color="FFFFFF", bold=True, size=10)
    normal_font  = Font(color="DDE6F0", size=9)
    thin = Side(style="thin", color="334466")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # ── Agrupar por equipamento ──
    equipamentos = df_filtrado["Equipamento"].unique()

    current_row = 1
    for equip in equipamentos:
        df_eq = df_filtrado[df_filtrado["Equipamento"] == equip].reset_index(drop=True)

        # Título
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=12)
        c = ws.cell(row=current_row, column=1, value="DIÁRIO DE BORDO")
        c.font = Font(color="FFFFFF", bold=True, size=14, name="Syne")
        c.fill = PatternFill("solid", fgColor="0D2137")
        c.alignment = center
        ws.row_dimensions[current_row].height = 28
        current_row += 1

        # Equipamento + Horímetro último abastec.
        try:
            hor_ult = df_eq["Horimetro_Ultimo_Abastec"].iloc[0] if not df_eq.empty else ""
        except (KeyError, IndexError):
            hor_ult = ""
        
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
        ws.cell(row=current_row, column=1, value="EQUIPAMENTO:").font  = white_font
        ws.cell(row=current_row, column=1).fill = subhead_fill
        ws.cell(row=current_row, column=1).alignment = center
        ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=7)
        ws.cell(row=current_row, column=3, value=equip).font = white_font
        ws.cell(row=current_row, column=3).fill = subhead_fill
        ws.cell(row=current_row, column=3).alignment = center
        ws.merge_cells(start_row=current_row, start_column=8, end_row=current_row, end_column=10)
        ws.cell(row=current_row, column=8, value="HORÍMETRO DO ÚLTIMO ABASTEC.").font = white_font
        ws.cell(row=current_row, column=8).fill = subhead_fill
        ws.cell(row=current_row, column=8).alignment = center
        ws.merge_cells(start_row=current_row, start_column=11, end_row=current_row, end_column=12)
        ws.cell(row=current_row, column=11, value=hor_ult).font = white_font
        ws.cell(row=current_row, column=11).fill = subhead_fill
        ws.cell(row=current_row, column=11).alignment = center
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Cabeçalho das colunas
        cabecalhos = [
            "DATA", "OPERADOR", "N° O.S.", "CLIENTE",
            "HORA\nINICIAL", "HORA\nFINAL",
            "HORÍMETRO\nINICIAL", "HORÍMETRO\nFINAL",
            "OBSERVAÇÕES",
            "HOR. TRAB.\nEQUIP.", "ABAST.\nHORÍMETRO", "MÉDIA"
        ]
        for col_idx, cab in enumerate(cabecalhos, start=1):
            c = ws.cell(row=current_row, column=col_idx, value=cab)
            c.font = white_font
            c.fill = header_fill
            c.alignment = center
            c.border = border
        ws.row_dimensions[current_row].height = 32
        current_row += 1

        # Dados
        for i, row in df_eq.iterrows():
            fill = alt_fill if i % 2 == 0 else PatternFill("solid", fgColor="131D28")
            valores = [
                row["Data"] if "Data" in row.index else "",
                row["Operador"] if "Operador" in row.index else "",
                row["N_OS"] if "N_OS" in row.index else "",
                row["Cliente"] if "Cliente" in row.index else "",
                row["Hora_Inicial"] if "Hora_Inicial" in row.index else "",
                row["Hora_Final"] if "Hora_Final" in row.index else "",
                row["Horimetro_Inicial"] if "Horimetro_Inicial" in row.index else "",
                row["Horimetro_Final"] if "Horimetro_Final" in row.index else "",
                row["Observacoes"] if "Observacoes" in row.index else "",
                row["Hor_Trab_Equip"] if "Hor_Trab_Equip" in row.index else "",
                row["Abast_Horimetro"] if "Abast_Horimetro" in row.index else "",
                row["Media"] if "Media" in row.index else "",
            ]
            aligns = [center, left, center, left, center, center,
                      center, center, left, center, center, center]
            for col_idx, (val, aln) in enumerate(zip(valores, aligns), start=1):
                c = ws.cell(row=current_row, column=col_idx, value=val)
                c.font = normal_font
                c.fill = fill
                c.alignment = aln
                c.border = border
            ws.row_dimensions[current_row].height = 18
            current_row += 1

        current_row += 2  # espaço entre equipamentos

    # Larguras das colunas
    larguras = [11, 18, 10, 20, 8, 8, 10, 10, 28, 10, 11, 8]
    for idx, larg in enumerate(larguras, start=1):
        ws.column_dimensions[ws.cell(1, idx).column_letter].width = larg

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────
#  CABEÇALHO
# ─────────────────────────────────────────────
st.markdown("""
<div class="header-box">
  <h1>⚙️ Diário de Bordo — Máquinas</h1>
  <p>Registro de operação de equipamentos · KAL · RS · EP</p>
</div>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
#  ABAS
# ─────────────────────────────────────────────
aba_form, aba_registros, aba_resumo = st.tabs([
    "📝  Novo Registro", "📋  Registros", "📊  Resumo"
])

# ═══════════════════════════════════════════
#  ABA 1 — FORMULÁRIO
# ═══════════════════════════════════════════
with aba_form:

    # — Equipamento —
    st.markdown("<div class='form-card'>", unsafe_allow_html=True)
    st.markdown("<div class='section-title'>Equipamento</div>", unsafe_allow_html=True)
    col_eq1, col_eq2 = st.columns(2)
    with col_eq1:
        equipamento = st.selectbox("🔧 Equipamento", EQUIPAMENTOS)
    with col_eq2:
        horimetro_ult_abastec = st.number_input(
            "⛽ Horímetro do Último Abastecimento",
            min_value=0.0, value=0.0, step=0.1, format="%.1f"
        )
    st.markdown("</div>", unsafe_allow_html=True)

    # — Identificação —
    st.markdown("<div class='form-card'>", unsafe_allow_html=True)
    st.markdown("<div class='section-title'>Identificação</div>", unsafe_allow_html=True)
    col_id1, col_id2, col_id3 = st.columns(3)
    with col_id1:
        data_reg = st.date_input("📅 Data", value=date.today())
    with col_id2:
        operador = st.text_input("👤 Operador", placeholder="Nome do operador")
    with col_id3:
        n_os = st.text_input("🗂️ N° O.S.", placeholder="Nº da Ordem de Serviço")
    cliente = st.text_input("🏢 Cliente", placeholder="Nome do cliente")
    st.markdown("</div>", unsafe_allow_html=True)

    # — Horários —
    st.markdown("<div class='form-card'>", unsafe_allow_html=True)
    st.markdown("<div class='section-title'>Horários de Operação</div>", unsafe_allow_html=True)
    col_h1, col_h2 = st.columns(2)
    with col_h1:
        hora_ini = st.time_input("🕐 Hora Inicial", value=time(7, 0))
    with col_h2:
        hora_fim = st.time_input("🕔 Hora Final", value=time(17, 0))

    horas_trabalhadas = calcular_horas(hora_ini, hora_fim)
    st.markdown(f"""
    <div class="calc-box">
      <div class="calc-item">
        <div class="calc-label">Horas Trabalhadas</div>
        <div class="calc-value">{horas_trabalhadas:.2f} h</div>
      </div>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("</div>", unsafe_allow_html=True)

    # — Horímetros —
    st.markdown("<div class='form-card'>", unsafe_allow_html=True)
    st.markdown("<div class='section-title'>Horímetros</div>", unsafe_allow_html=True)
    col_km1, col_km2 = st.columns(2)
    with col_km1:
        hor_ini = st.number_input("📟 Horímetro Inicial", min_value=0.0, value=0.0, step=0.1, format="%.1f")
    with col_km2:
        hor_fim = st.number_input("📟 Horímetro Final", min_value=0.0, value=0.0, step=0.1, format="%.1f")

    hor_trab = round(hor_fim - hor_ini, 2)
    abast_hor = round(hor_fim - horimetro_ult_abastec, 2) if horimetro_ult_abastec > 0 else 0.0
    media = round(abast_hor / horas_trabalhadas, 2) if horas_trabalhadas > 0 and abast_hor > 0 else 0.0

    if hor_trab < 0:
        st.error("⚠️ Horímetro Final não pode ser menor que o Inicial.")
    else:
        st.markdown(f"""
        <div class="calc-box">
          <div class="calc-item">
            <div class="calc-label">Hor. Trab. Equip.</div>
            <div class="calc-value">{hor_trab:.2f} h</div>
          </div>
          <div class="calc-item">
            <div class="calc-label">Abast. Horímetro</div>
            <div class="calc-value">{abast_hor:.2f}</div>
          </div>
          <div class="calc-item">
            <div class="calc-label">Média</div>
            <div class="calc-value">{media:.2f}</div>
          </div>
        </div>
        """, unsafe_allow_html=True)
    st.markdown("</div>", unsafe_allow_html=True)

    # — Observações —
    st.markdown("<div class='form-card'>", unsafe_allow_html=True)
    st.markdown("<div class='section-title'>Observações</div>", unsafe_allow_html=True)
    observacoes = st.text_area(
        "📝 Observações",
        placeholder="Descreva ocorrências, manutenções, paradas, anomalias...",
        height=100
    )
    st.markdown("</div>", unsafe_allow_html=True)

    # Botão salvar
    if st.button("💾  Salvar Registro", type="primary"):
        erros = []
        if not operador:
            erros.append("Operador")
        if not cliente:
            erros.append("Cliente")
        if hor_trab < 0:
            erros.append("Horímetro (Final < Inicial)")
        if erros:
            st.error(f"⚠️ Corrija os campos: {', '.join(erros)}")
        else:
            novo_id = f"RM-{datetime.now().strftime('%Y%m%d%H%M%S')}"
            registro = {
                "ID": novo_id,
                "Equipamento": equipamento,
                "Horimetro_Ultimo_Abastec": horimetro_ult_abastec,
                "Data": str(data_reg),
                "Operador": operador,
                "N_OS": n_os,
                "Cliente": cliente,
                "Hora_Inicial": str(hora_ini),
                "Hora_Final": str(hora_fim),
                "Horimetro_Inicial": hor_ini,
                "Horimetro_Final": hor_fim,
                "Hor_Trab_Equip": horas_trabalhadas,
                "Abast_Horimetro": abast_hor,
                "Media": media,
                "Observacoes": observacoes,
                "Registro_Em": datetime.now().strftime("%d/%m/%Y %H:%M")
            }
            salvar_registro(registro)
            st.success(f"✅ Registro **{novo_id}** salvo com sucesso! Equipamento: **{equipamento}**")
            st.balloons()

# ═══════════════════════════════════════════
#  ABA 2 — REGISTROS
# ═══════════════════════════════════════════
with aba_registros:
    df = carregar_dados()

    if df.empty:
        st.info("Nenhum registro ainda. Preencha o formulário na aba **Novo Registro**!")
    else:
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            equip_filtro = st.selectbox("🔧 Filtrar por equipamento", ["Todos"] + EQUIPAMENTOS)
        with col_f2:
            busca_op = st.text_input("🔍 Buscar por operador ou cliente", "")

        df_filtrado = df.copy()
        if equip_filtro != "Todos":
            df_filtrado = df_filtrado[df_filtrado["Equipamento"] == equip_filtro]
        if busca_op:
            mask = (
                df_filtrado["Operador"].str.contains(busca_op, case=False, na=False) |
                df_filtrado["Cliente"].str.contains(busca_op, case=False, na=False)
            )
            df_filtrado = df_filtrado[mask]

        st.markdown(f"**{len(df_filtrado)} registro(s) encontrado(s)**")
        st.dataframe(
            df_filtrado[[
                "Data", "Equipamento", "Operador", "Cliente",
                "N_OS", "Hora_Inicial", "Hora_Final",
                "Hor_Trab_Equip", "Abast_Horimetro", "Media"
            ]].rename(columns={
                "Hor_Trab_Equip": "H. Trabalhadas",
                "Abast_Horimetro": "Abast. Hor.",
                "N_OS": "N° O.S."
            }),
            use_container_width=True,
            hide_index=True
        )

        col_dl1, col_dl2 = st.columns(2)
        with col_dl1:
            csv_bytes = df_filtrado.to_csv(index=False).encode("utf-8")
            st.download_button("⬇️  Baixar CSV", data=csv_bytes,
                               file_name="diario_maquinas.csv", mime="text/csv")
        with col_dl2:
            excel_bytes = gerar_excel(df_filtrado)
            st.download_button("📄  Exportar Excel (.xlsx)", data=excel_bytes,
                               file_name="diario_maquinas.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ═══════════════════════════════════════════
#  ABA 3 — RESUMO
# ═══════════════════════════════════════════
with aba_resumo:
    df = carregar_dados()

    if df.empty:
        st.info("Ainda sem dados para o resumo.")
    else:
        total_reg     = len(df)
        total_horas   = df["Hor_Trab_Equip"].sum()
        total_equip   = df["Equipamento"].nunique()
        media_global  = df["Media"].replace(0, pd.NA).mean()

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Total de Registros", total_reg)
        c2.metric("Horas Operadas", f"{total_horas:.1f} h")
        c3.metric("Equipamentos", total_equip)
        c4.metric("Média Global", f"{media_global:.2f}" if pd.notna(media_global) else "—")

        st.markdown("---")

        st.markdown("**⏱️ Horas por Equipamento**")
        hor_equip = (
            df.groupby("Equipamento")["Hor_Trab_Equip"]
            .sum().sort_values(ascending=False).reset_index()
        )
        hor_equip.columns = ["Equipamento", "Horas"]
        st.bar_chart(hor_equip.set_index("Equipamento"))

        st.markdown("**👤 Horas por Operador**")
        hor_op = (
            df.groupby("Operador")["Hor_Trab_Equip"]
            .sum().sort_values(ascending=False).reset_index()
        )
        hor_op.columns = ["Operador", "Horas"]
        st.bar_chart(hor_op.set_index("Operador"))

# ─────────────────────────────────────────────
#  RODAPÉ
# ─────────────────────────────────────────────
st.markdown("""
<div class="footer">
  Diário de Bordo · Máquinas e Equipamentos · Dados salvos em CSV local
</div>
""", unsafe_allow_html=True)